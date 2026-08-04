"""Session spreadsheet derivatives (CSV / XLSX).

Authoritative truth remains session JSONL. CSV/XLSX are rebuildable handoffs.

Operator-facing columns:
- cultivator ← facility_id (company / grower)
- strain ← cultivar_normalized_name (sticky strain)

Internal JSONL still uses cultivar_* for strain and facility_id for cultivator.
"""
from __future__ import annotations

import csv
import os
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

HEADERS = [
    "sequence",
    "record_id",
    "captured_at",
    "barcode_raw",
    "barcode_normalized",
    "cultivator",
    "strain",
    "cultivar_raw_name",
    "cultivar_normalized_name",
    "run_id",
    "container_id",
    "tare_g",
    "gross_g",
    "net_g",
    "operator_id",
    "station_id",
    "device_id",
    "calibration_id",
    "capture_mode",
    "duplicate_status",
    "record_hash",
    "operator_note",
    "void_status",
]


def safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def row_for(record: dict[str, Any]) -> list[Any]:
    """Build one spreadsheet row; derive cultivator/strain from stable JSONL keys."""
    strain = record.get("cultivar_normalized_name") or record.get("cultivar_raw_name") or ""
    cultivator = record.get("facility_id") or record.get("cultivator") or ""
    enriched = dict(record)
    enriched.setdefault("strain", strain)
    enriched.setdefault("cultivator", cultivator)
    return [safe_cell(enriched.get(h, "")) for h in HEADERS]


def append_csv(path: Path, record: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    new = not path.exists()
    if not new:
        with path.open("r", newline="", encoding="utf-8") as handle:
            existing = next(csv.reader(handle), [])
        if existing[: len(HEADERS)] != HEADERS:
            backup = path.with_suffix(path.suffix + ".incompatible.backup")
            if not backup.exists():
                shutil.copy2(path, backup)
            raise ValueError("incompatible spreadsheet header")
    with path.open("a", newline="", encoding="utf-8") as handle:
        out = csv.writer(handle)
        if new:
            out.writerow(HEADERS)
        out.writerow(row_for(record))
        handle.flush()
        os.fsync(handle.fileno())


def append_xlsx(path: Path, record: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        if existing[: len(HEADERS)] != HEADERS:
            backup = path.with_suffix(path.suffix + ".incompatible.backup")
            if not backup.exists():
                shutil.copy2(path, backup)
            raise ValueError("incompatible spreadsheet header")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Weights"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(ord('A') + len(HEADERS) - 1)}1"
    ws.append(row_for(record))
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, path)


def rebuild_spreadsheets_from_jsonl(session_dir: Path, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Rebuild records.csv / records.xlsx from accepted JSONL weight records.

    JSONL remains authoritative. Existing CSV/XLSX are replaced atomically after
    a backup of any prior file.
    """
    accepted = [
        row
        for row in rows
        if row.get("event_type") == "weight_record" and row.get("record_status") == "accepted"
    ]
    accepted.sort(key=lambda row: int(row.get("sequence", 0)))
    csv_path = session_dir / "records.csv"
    xlsx_path = session_dir / "records.xlsx"
    for path in (csv_path, xlsx_path):
        if path.exists():
            backup = path.with_suffix(path.suffix + ".pre_rebuild.backup")
            shutil.copy2(path, backup)
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        out = csv.writer(handle)
        out.writerow(HEADERS)
        for row in accepted:
            out.writerow(row_for(row))
        handle.flush()
        os.fsync(handle.fileno())
    wb = Workbook()
    ws = wb.active
    ws.title = "Weights"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    for row in accepted:
        ws.append(row_for(row))
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, xlsx_path)
    return {
        "rebuilt_rows": len(accepted),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
    }
