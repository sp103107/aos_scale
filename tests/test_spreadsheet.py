from openpyxl import load_workbook

from best_buds_weight_station.spreadsheet import HEADERS, append_csv, append_xlsx, row_for
from best_buds_weight_station.run_manager import facility_id_from_cultivator


def test_formula_injection(tmp_path):
    r = {h: "" for h in HEADERS}
    r.update({"sequence": 1, "record_id": "1", "barcode_raw": '=HYPERLINK("x")'})
    append_xlsx(tmp_path / "a.xlsx", r)
    ws = load_workbook(tmp_path / "a.xlsx").active
    assert str(ws.cell(2, 4).value).startswith("'=")
    append_csv(tmp_path / "a.csv", r)
    assert "'=HYPERLINK" in (tmp_path / "a.csv").read_text()


def test_compatible_continuation(tmp_path):
    r = {h: h for h in HEADERS}
    append_xlsx(tmp_path / "a.xlsx", r)
    append_xlsx(tmp_path / "a.xlsx", r)
    assert load_workbook(tmp_path / "a.xlsx").active.max_row == 3


def test_row_for_derives_cultivator_and_strain():
    record = {
        "sequence": 1,
        "record_id": "r1",
        "captured_at": "2026-08-04T00:00:00Z",
        "barcode_raw": "P-1",
        "barcode_normalized": "P-1",
        "facility_id": "Best-Buds",
        "cultivar_raw_name": "Blue Dream",
        "cultivar_normalized_name": "Blue Dream",
        "run_id": "RUN-1",
        "container_id": "DEFAULT",
        "tare_g": 0.0,
        "gross_g": 100.0,
        "net_g": 100.0,
        "operator_id": "op",
        "station_id": "WEIGHT-STATION-01",
    }
    values = dict(zip(HEADERS, row_for(record)))
    assert "cultivator" in HEADERS
    assert "strain" in HEADERS
    assert values["cultivator"] == "Best-Buds"
    assert values["strain"] == "Blue Dream"
    assert values["cultivar_normalized_name"] == "Blue Dream"


def test_facility_id_from_cultivator_slug():
    assert facility_id_from_cultivator("Best Buds") == "Best-Buds"
    assert facility_id_from_cultivator("") == "BEST-BUDS"
