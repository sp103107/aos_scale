"""BBWS SR5 artifact polish contract tests — handoff only; JSONL authoritative."""
from pathlib import Path

from openpyxl import load_workbook

from best_buds_weight_station.models import CaptureCommand, RunContext
from best_buds_weight_station.reports import NON_CLAIMS, compile_report, reconcile_export_to_jsonl
from best_buds_weight_station.spreadsheet import HEADERS
from best_buds_weight_station.storage import SessionStore


ROOT = Path(__file__).parents[1]


def _session(tmp_path: Path) -> SessionStore:
    ctx = RunContext("S1", "RUN-SR5", "OP1", "BestBuds", "ST1", "C1", "Blue Dream", "Blue Dream", "BIN", 1)
    store = SessionStore(tmp_path, ctx)
    store.commit(CaptureCommand("TAG-001", 11.0, 8.0, {}, "manual"))
    store.commit(CaptureCommand("TAG-002", 12.5, 9.5, {}, "manual"))
    return store


def test_compile_keeps_cultivator_strain_headers(tmp_path):
    store = _session(tmp_path)
    report = compile_report(store.session_dir)
    plants = Path(report["artifacts"]["csv_plants"])
    header = plants.read_text(encoding="utf-8").splitlines()[0].split(",")
    assert header == HEADERS
    assert "cultivator" in header and "strain" in header


def test_summary_csv_includes_cultivator_context(tmp_path):
    store = _session(tmp_path)
    report = compile_report(store.session_dir)
    text = Path(report["artifacts"]["csv_summary"]).read_text(encoding="utf-8")
    assert text.startswith("cultivator,strain,net_g")
    assert "Blue Dream" in text
    assert "TOTAL" in text


def test_xlsx_has_nonclaims_and_styled_sheets(tmp_path):
    store = _session(tmp_path)
    report = compile_report(store.session_dir)
    wb = load_workbook(report["artifacts"]["xlsx"])
    assert set(wb.sheetnames) >= {"Summary", "Records", "NonClaims"}
    records = wb["Records"]
    assert [c.value for c in records[1]] == list(HEADERS)
    claims = wb["NonClaims"]
    body = " ".join(str(c.value or "") for row in claims.iter_rows(min_row=2) for c in row)
    assert "JSONL" in body or "jsonl" in body.lower()
    assert "Metrc" in body or "legal-for-trade" in body.lower() or "metrology" in body.lower()


def test_handoff_bundle_manifest(tmp_path):
    store = _session(tmp_path)
    report = compile_report(store.session_dir)
    bundle_path = Path(report["artifacts"]["handoff_bundle_manifest"])
    assert bundle_path.exists()
    import json

    bundle = json.loads(bundle_path.read_text(encoding="utf-8"))
    assert bundle["manifest_type"] == "bbws.handoff_bundle"
    assert bundle["authoritative"] is False
    assert bundle["records_sha256"] == report["records_sha256"]
    assert "xlsx" in bundle["artifacts"] and "docx" in bundle["artifacts"]
    assert bundle["non_claims"] == list(NON_CLAIMS)


def test_docx_contains_non_claim_and_cultivator_label(tmp_path):
    store = _session(tmp_path)
    report = compile_report(store.session_dir)
    from docx import Document

    doc = Document(report["artifacts"]["docx"])
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "non-authoritative" in text.lower()
    assert "Cultivator" in text
    assert "JSONL" in text or "jsonl" in text.lower()


def test_reconcile_pass_includes_bundle(tmp_path):
    store = _session(tmp_path)
    receipt = reconcile_export_to_jsonl(store.session_dir)
    assert receipt["status"] == "pass"
    assert receipt["headers_ok"] is True
    assert receipt["bundle_ok"] is True
    assert receipt["handoff_bundle_manifest"]
    assert receipt["non_claims"]


def test_selection_map_and_runbook_exist():
    assert (ROOT / "docs" / "BBWS_SR5_SELECTION_MAP.md").exists()
    assert (ROOT / "docs" / "BBWS_SR5_ARTIFACT_POLISH_RUNBOOK.md").exists()
